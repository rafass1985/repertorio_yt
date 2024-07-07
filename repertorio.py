import pandas as pd
from youtube_search import YoutubeSearch
import json
import openpyxl
import os
from openpyxl.styles import Font

# Criar uma planilha para os resultados
book = openpyxl.Workbook()
print(book.sheetnames)
videos_page = book['Sheet']

# Caminho para o arquivo Excel
file_path = "repertorio.xlsx"

# Nome da planilha
sheet_name = "repertorio"

# Leitura dos dados da planilha Excel
df = pd.read_excel(file_path, sheet_name=sheet_name, header=None)

# Acessar a coluna A
coluna_A = df[0]
##print(df)

# Adicionar cabeçalho na planilha
videos_page.append(["Original", "Título", "Duração", "URL"])

# Imprimir cada item da coluna 1 separadamente
for item in coluna_A:
    results = YoutubeSearch(item, max_results=1).to_json()
    dados = json.loads(results)
    
    if dados['videos']:
        video_info = dados['videos'][0]
        title = video_info['title']
        duration = video_info['duration']
        url = "https://www.youtube.com" + video_info['url_suffix']
        
        # Adiciona os dados na planilha
        row = [item, title, duration, url]
        videos_page.append(row)

        # Gerar a URL do video
        cell = videos_page.cell(row=videos_page.max_row, column=4)
        cell.hyperlink = url
        cell.font = Font(color="0000FF", underline="single")
        
        # Firula pro usuario acompanhar
        print("\nBuscando: " + item)
        print("Encontrado: " + title)
    else:
        # Caso não encontre resultados
        row = [item, "Nenhum resultado encontrado", "-", "-"]
        videos_page.append(row)
        
        # Firula pro usuario acompanhar
        print("\nBuscando: " + item)
        print("Nenhum resultado encontrado")

# Salvar o arquivo Excel com a lista de vídeos
output_path = os.path.join(os.path.dirname(file_path), 'lista_videos.xlsx')
book.save(output_path)